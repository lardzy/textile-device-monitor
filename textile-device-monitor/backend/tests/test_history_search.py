from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker

from app import database as database_module
from app.api.history import export_history, get_history
from app.crud import history as history_crud
from app.models import Base, Device, DeviceStatusHistory
from app.utils.exporters import export_history_to_excel


class HistorySearchTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:")
        cls.SessionLocal = sessionmaker(bind=cls.engine)

    @classmethod
    def tearDownClass(cls):
        cls.engine.dispose()

    def setUp(self):
        Base.metadata.create_all(bind=self.engine)
        self.db = self.SessionLocal()
        device = Device(device_code="history-device", name="历史设备")
        self.db.add(device)
        self.db.commit()
        self.db.refresh(device)
        self.device = device

        self._add_history("TASK-Alpha-01", "Cotton Sample")
        self._add_history("task-beta-02", "Wool Inspection")
        self._add_history("literal-100%_done", "Special Pattern")

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(bind=self.engine)

    def _add_history(
        self,
        task_id: str,
        task_name: str,
        inspector_name: str | None = None,
    ) -> None:
        self.db.add(
            DeviceStatusHistory(
                device_id=self.device.id,
                status="idle",
                task_id=task_id,
                task_name=task_name,
                inspector_name=inspector_name,
                task_progress=100,
                reported_at=datetime(2026, 7, 1, 0, 0, tzinfo=timezone.utc),
            )
        )
        self.db.commit()

    def test_keyword_matches_task_id_or_name_case_insensitively(self):
        by_id, id_total = history_crud.get_device_history(
            self.db,
            keyword="alpha",
            limit=20,
        )
        by_name, name_total = history_crud.get_device_history(
            self.db,
            keyword="WOOL",
            limit=20,
        )

        self.assertEqual(id_total, 1)
        self.assertEqual(by_id[0].task_id, "TASK-Alpha-01")
        self.assertEqual(name_total, 1)
        self.assertEqual(by_name[0].task_id, "task-beta-02")

    def test_keyword_treats_like_wildcards_as_literal_text(self):
        history, total = history_crud.get_device_history(
            self.db,
            keyword="%_",
            limit=20,
        )

        self.assertEqual(total, 1)
        self.assertEqual(history[0].task_id, "literal-100%_done")

    def test_keyword_matches_persisted_queue_person(self):
        self._add_history("task-person", "Person Search", "张三")

        history, total = history_crud.get_device_history(
            self.db,
            keyword="张三",
            limit=20,
        )

        self.assertEqual(total, 1)
        self.assertEqual(history[0].inspector_name, "张三")

    def test_list_and_excel_include_queue_person_snapshot(self):
        self._add_history("task-export", "Export Task", "李工")

        response = get_history(
            device_id=self.device.id,
            start_date=None,
            end_date=None,
            status=None,
            task_id="task-export",
            keyword=None,
            page=1,
            page_size=20,
            db=self.db,
        )
        self.assertEqual(response["data"][0]["inspector_name"], "李工")

        history_record = (
            self.db.query(DeviceStatusHistory)
            .filter(DeviceStatusHistory.task_id == "task-export")
            .one()
        )
        excel_response = export_history_to_excel([history_record])
        workbook = load_workbook(BytesIO(excel_response.body), read_only=True)
        worksheet = workbook["设备状态历史"]
        headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        values = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
        exported = dict(zip(headers, values))
        self.assertEqual(exported["排队人员"], "李工")

    def test_excel_escapes_formula_like_queue_person(self):
        self._add_history("task-safe-export", "Safe Export", "=1+1")
        history_record = (
            self.db.query(DeviceStatusHistory)
            .filter(DeviceStatusHistory.task_id == "task-safe-export")
            .one()
        )

        excel_response = export_history_to_excel([history_record])
        workbook = load_workbook(BytesIO(excel_response.body), read_only=True)
        worksheet = workbook["设备状态历史"]
        headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        values = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
        exported = dict(zip(headers, values))
        self.assertEqual(exported["排队人员"], "'=1+1")

    def test_compatibility_schema_adds_queue_person_column(self):
        legacy_engine = create_engine("sqlite:///:memory:")
        try:
            with legacy_engine.begin() as connection:
                connection.execute(
                    text(
                        "CREATE TABLE device_status_history ("
                        "id INTEGER PRIMARY KEY, device_id INTEGER NOT NULL, "
                        "status VARCHAR(20) NOT NULL)"
                    )
                )

            with patch.object(database_module, "engine", legacy_engine):
                database_module.ensure_device_status_history_schema()
                database_module.ensure_device_status_history_schema()

            columns = {
                column["name"]
                for column in inspect(legacy_engine).get_columns(
                    "device_status_history"
                )
            }
            self.assertIn("inspector_name", columns)
        finally:
            legacy_engine.dispose()

    def test_end_date_is_exclusive(self):
        boundary = datetime(2026, 7, 2, 0, 0, tzinfo=timezone.utc)
        self._add_history("before-boundary", "Before Boundary")
        before = (
            self.db.query(DeviceStatusHistory)
            .filter(DeviceStatusHistory.task_id == "before-boundary")
            .one()
        )
        before.reported_at = boundary - timedelta(microseconds=1)
        self.db.add(
            DeviceStatusHistory(
                device_id=self.device.id,
                status="idle",
                task_id="at-boundary",
                task_name="At Boundary",
                task_progress=100,
                reported_at=boundary,
            )
        )
        self.db.commit()

        history, _ = history_crud.get_device_history(
            self.db,
            start_date=datetime(2026, 7, 1, tzinfo=timezone.utc),
            end_date=boundary,
            limit=20,
        )

        task_ids = {item.task_id for item in history}
        self.assertIn("before-boundary", task_ids)
        self.assertNotIn("at-boundary", task_ids)

    def test_list_and_export_forward_the_same_keyword(self):
        with patch(
            "app.api.history.history_crud.get_device_history",
            return_value=([], 0),
        ) as list_query:
            get_history(
                device_id=None,
                start_date=None,
                end_date=None,
                status=None,
                task_id=None,
                keyword="needle",
                page=1,
                page_size=20,
                db=self.db,
            )
        self.assertEqual(list_query.call_args.kwargs["keyword"], "needle")

        with (
            patch(
                "app.api.history.history_crud.get_device_history",
                return_value=([object()], 1),
            ) as export_query,
            patch(
                "app.api.history.export_history_to_excel",
                return_value="export-response",
            ),
        ):
            result = export_history(
                device_id=None,
                start_date=None,
                end_date=None,
                status=None,
                task_id=None,
                keyword="needle",
                db=self.db,
            )

        self.assertEqual(result, "export-response")
        self.assertEqual(export_query.call_args.kwargs["keyword"], "needle")


if __name__ == "__main__":
    unittest.main()
